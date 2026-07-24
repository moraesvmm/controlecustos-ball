from fastapi import FastAPI, Request, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
import os
import sqlite3
import json
import asyncio
import tempfile
import shutil
try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

ROOT_DIR = os.path.join(os.path.dirname(__file__), "..")
DB_PATH = os.path.join(os.path.dirname(__file__), "database", "database.sqlite")
SYNC_PATH = os.path.join(os.path.dirname(__file__), "database", "database.sync")

# ==============================================================================
# ARQUITETURA DISTRIBUÍDA (Multi-Servidor / Banco Único)
# ==============================================================================
# Responsabilidade deste Backend: Servir de intermediário local para o navegador
# do usuário, processando regras de negócio e gravando no arquivo SQLite.
#
# Atenção: Como este servidor Python roda localmente na máquina de CADA usuário
# (ou seja, não há um servidor central), a comunicação "Real-Time" (SSE) 
# funciona monitorando a data de modificação física do arquivo `database.sqlite` 
# na rede compartilhada. Se a data (mtime) muda, significa que outro PC salvou
# dados, e este servidor avisa a tela local (frontend) para recarregar.
# ==============================================================================

app = FastAPI(title="Controle RC Backend (Localhost SQLite)")

@app.middleware("http")
async def sse_trigger_middleware(request: Request, call_next):
    override = request.headers.get("x-http-method-override", "").upper()
    method = override if override else request.method
    path = request.url.path
    
    response = await call_next(request)
    
    # Touch the sync file if it's a modifying request
    if method in ["POST", "PATCH", "PUT", "DELETE"] and ("/rest/v1/" in path or "/api/" in path):
        if 200 <= response.status_code < 300:
            if path != "/api/stream":
                try:
                    with open(SYNC_PATH, 'a'):
                        os.utime(SYNC_PATH, None)
                except Exception:
                    pass
    return response

@app.get("/api/custo-geral/movimentacoes")
def get_custo_geral_movimentacoes():
    try:
        conn = get_db()
        
        # Lê as movimentacoes reais da tabela SQLite
        res_rows = conn.execute("SELECT * FROM custo_geral").fetchall()
        res = [dict(r) for r in res_rows]
        
        covered_months = set()
        for d in res:
            if d.get("mes"):
                covered_months.add(str(int(d["mes"])))
        
        # Lê os agregados historicos dos meses passados
        rows = conn.execute("SELECT mes, manutencao, ferramentaria, facilities FROM custo_geral_mensal").fetchall()
        
        mes_map = {'jan': '1', 'fev': '2', 'mar': '3', 'abr': '4', 'mai': '5', 'jun': '6', 'jul': '7', 'ago': '8', 'set': '9', 'out': '10', 'nov': '11', 'dez': '12'}
        
        for r in rows:
            m_str = mes_map.get(str(r['mes']).lower(), '1')
            
            # Pular se o mês atual já tem dados reais importados
            if m_str in covered_months:
                continue
                
            # Adicionar agregados historicos
            if r['manutencao'] and r['manutencao'] > 0:
                res.append({'mes': m_str, 'custo_do_mes': r['manutencao'], 'check': 'MANUTENÇÃO - Real Consumo', 'descricao_conta': 'MANUTENCAO', 'grupo': 'Histórico', 'area': 'MANUTENÇÃO', 'dt_trans': f"2026-{int(m_str):02d}-01"})
            if r['ferramentaria'] and r['ferramentaria'] > 0:
                res.append({'mes': m_str, 'custo_do_mes': r['ferramentaria'], 'check': 'FERRAMENTARIA - Real Consumo', 'descricao_conta': 'FERRAMENTAS', 'grupo': 'Histórico', 'area': 'FERRAMENTARIA', 'dt_trans': f"2026-{int(m_str):02d}-01"})
            if r['facilities'] and r['facilities'] > 0:
                res.append({'mes': m_str, 'custo_do_mes': r['facilities'], 'check': 'FACILITIES - Real Consumo', 'descricao_conta': 'FACILITIES', 'grupo': 'Histórico', 'area': 'FACILITIES', 'dt_trans': f"2026-{int(m_str):02d}-01"})
                
        return res
    except Exception as e:
        print(f"Erro ao buscar movimentacoes: {e}")
        return [{"error": str(e)}]


# Caminho onde guardamos o ultimo arquivo xlsm enviado (por maquina)
EXCEL_CACHE_PATH = os.path.join(os.path.dirname(__file__), "database", "last_upload.xlsm")
HISTORICO_DIR = r"C:\Users\VMORAES1\Documents\A ALIMENTAR\2026"

def _parse_excel(filepath: str) -> dict:
    """Le o arquivo xlsm e extrai dashboard + movimentacoes."""
    if not OPENPYXL_OK:
        return {"error": "openpyxl nao instalado. Execute: pip install openpyxl"}
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        result = {"dashboard": {}, "movimentacoes": [], "sheetNames": wb.sheetnames}

        # ---- ABA DASHBOARD ----
        dash_sheet = None
        for name in wb.sheetnames:
            if "dashboard" in name.lower():
                dash_sheet = wb[name]
                break

        if dash_sheet:
            # Varrer todas as celulas procurando valores-chave
            # Estrategia: procurar labels de texto e pegar o valor da celula adjacente (direita ou baixo)
            dashboard_data = {}
            meses = ['jan','fev','mar','abr','mai','jun','jul','ago','set','out','nov','dez']
            mes_map = {m: i for i, m in enumerate(meses)}

            # Acumular dados mensais por area
            evolucao = {"manutencao": [0]*12, "ferramentaria": [0]*12, "facilities": [0]*12, "budget": [0]*12}

            for row in dash_sheet.iter_rows(values_only=True):
                row_vals = [str(v).strip() if v is not None else '' for v in row]
                row_str = ' '.join(row_vals).lower()

                # Procurar budget total anual
                for i, cell in enumerate(row):
                    if cell is None:
                        continue
                    cell_str = str(cell).lower().strip()

                    # Budget / Orcamento
                    if any(k in cell_str for k in ['budget', 'orçamento', 'orcamento', 'meta anual']) and i + 1 < len(row):
                        next_val = row[i+1]
                        if isinstance(next_val, (int, float)) and next_val > 1000:
                            dashboard_data.setdefault('budget_anual', float(next_val))

                    # Realizado por area
                    if 'manutenção' in cell_str or 'manutencao' in cell_str:
                        for j in range(i+1, min(i+14, len(row))):
                            v = row[j]
                            if isinstance(v, (int, float)) and v != 0:
                                col_idx = j - i - 1
                                if 0 <= col_idx < 12:
                                    evolucao['manutencao'][col_idx] += abs(float(v))

                    if 'ferramentaria' in cell_str:
                        for j in range(i+1, min(i+14, len(row))):
                            v = row[j]
                            if isinstance(v, (int, float)) and v != 0:
                                col_idx = j - i - 1
                                if 0 <= col_idx < 12:
                                    evolucao['ferramentaria'][col_idx] += abs(float(v))

                    if 'facilities' in cell_str or 'facilidades' in cell_str:
                        for j in range(i+1, min(i+14, len(row))):
                            v = row[j]
                            if isinstance(v, (int, float)) and v != 0:
                                col_idx = j - i - 1
                                if 0 <= col_idx < 12:
                                    evolucao['facilities'][col_idx] += abs(float(v))

                    if 'budget' in cell_str and 'mensal' in cell_str:
                        for j in range(i+1, min(i+14, len(row))):
                            v = row[j]
                            if isinstance(v, (int, float)) and v > 0:
                                col_idx = j - i - 1
                                if 0 <= col_idx < 12:
                                    evolucao['budget'][col_idx] = float(v)

            dashboard_data['evolucao'] = evolucao

            # Extrair KPIs principais: realizado total por area (soma dos meses ja realizados)
            hoje_mes = __import__('datetime').datetime.now().month - 1  # 0-indexed
            dashboard_data['realizado_manutencao'] = sum(evolucao['manutencao'][:hoje_mes+1])
            dashboard_data['realizado_ferramentaria'] = sum(evolucao['ferramentaria'][:hoje_mes+1])
            dashboard_data['realizado_facilities'] = sum(evolucao['facilities'][:hoje_mes+1])
            dashboard_data['realizado_total'] = (
                dashboard_data['realizado_manutencao'] +
                dashboard_data['realizado_ferramentaria'] +
                dashboard_data['realizado_facilities']
            )

            result['dashboard'] = dashboard_data

        # ---- ABA MOVIMENTACOES ----
        mov_sheet = None
        for name in wb.sheetnames:
            n = name.lower()
            if 'movimenta' in n or 'custo geral' in n:
                mov_sheet = wb[name]
                break

        if mov_sheet:
            rows_iter = mov_sheet.iter_rows(values_only=True)
            header = None
            records = []
            for row in rows_iter:
                if header is None:
                    # Primeiro row nao-vazio com pelo menos 3 celulas preenchidas = header
                    non_null = [c for c in row if c is not None and str(c).strip() != '']
                    if len(non_null) >= 3:
                        header = [str(c).strip().lower().replace(' ', '_').replace('-','_') if c else f'col{i}'
                                  for i, c in enumerate(row)]
                    continue
                if all(c is None or str(c).strip() == '' for c in row):
                    continue
                rec = {header[i]: (row[i] if i < len(row) else None) for i in range(len(header))}
                # Converter datas
                for k, v in rec.items():
                    if hasattr(v, 'strftime'):
                        rec[k] = v.strftime('%Y-%m-%d')
                    elif isinstance(v, float) and v == int(v) and 'custo' not in k and 'valor' not in k and 'material' not in k:
                        rec[k] = int(v)
                records.append(rec)

            result['movimentacoes'] = records

        wb.close()
        return result
    except Exception as e:
        import traceback
        return {"error": str(e), "trace": traceback.format_exc()}


@app.get("/api/custo-geral/movimentacoes")
def get_custo_geral_movimentacoes():
    """Releitura do ultimo Excel enviado — retorna dados do dashboard + movimentacoes."""
    if os.path.exists(EXCEL_CACHE_PATH):
        return _parse_excel(EXCEL_CACHE_PATH)
    # Fallback: banco SQLite legado
    try:
        conn = get_db()
        res_rows = conn.execute("SELECT * FROM custo_geral LIMIT 5000").fetchall()
        return {"movimentacoes": [dict(r) for r in res_rows], "dashboard": {}, "source": "sqlite_fallback"}
    except Exception as e:
        return {"movimentacoes": [], "dashboard": {}, "error": str(e)}




@app.get("/api/custo-geral/historico")
def get_custo_geral_historico():
    """Varre a pasta A ALIMENTAR\2026 e consolida os dados mensais historicos."""
    if not OPENPYXL_OK:
        return {"error": "openpyxl nao instalado"}
    if not os.path.exists(HISTORICO_DIR):
        return {"meses": [], "error": f"Pasta nao encontrada: {HISTORICO_DIR}"}
    consolidado = {}  # {mes_idx: {manutencao, ferramentaria, facilities, budget}}
    try:
        for root, dirs, files in os.walk(HISTORICO_DIR):
            for fname in files:
                if not fname.endswith('.xlsm') and not fname.endswith('.xlsx'):
                    continue
                if fname.startswith('~'):
                    continue
                fpath = os.path.join(root, fname)
                try:
                    parsed = _parse_excel(fpath)
                    ev = parsed.get('dashboard', {}).get('evolucao', {})
                    for i in range(12):
                        if i not in consolidado:
                            consolidado[i] = {'manutencao': 0, 'ferramentaria': 0, 'facilities': 0, 'budget': 0}
                        consolidado[i]['manutencao'] += ev.get('manutencao', [0]*12)[i]
                        consolidado[i]['ferramentaria'] += ev.get('ferramentaria', [0]*12)[i]
                        consolidado[i]['facilities'] += ev.get('facilities', [0]*12)[i]
                        b = ev.get('budget', [0]*12)[i]
                        if b > 0:
                            consolidado[i]['budget'] = b
                except Exception:
                    continue
        meses_label = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
        return {"meses": [{"mes": meses_label[i], "mes_idx": i, **consolidado.get(i, {})} for i in range(12)]}
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/custo-geral/mensal")
def get_custo_geral_mensal(ano: int = 2026):
    try:
        conn = get_db()
        rows = conn.execute("SELECT * FROM custo_geral_mensal WHERE ano=? ORDER BY id", (ano,)).fetchall()
        return [dict(r) for r in rows]
    except Exception as e:
        return []

# ==============================================================================
# NOVO MÓDULO: MOVIMENTAÇÕES (Inteligência Financeira)
# ==============================================================================

@app.post("/api/movimentacoes/import")
async def import_movimentacoes_post(file: UploadFile = File(...)):
    if not OPENPYXL_OK:
        raise HTTPException(status_code=400, detail="Pandas/Openpyxl não estão instalados no servidor.")
    
    import pandas as pd
    from io import BytesIO
    content = await file.read()
    
    try:
        # Puxa o mês e ano do nome do arquivo ou da aba (aqui assumiremos 2026 e mes fixo para o teste se não houver)
        # Tenta achar o mês a partir da aba DADOS, que tem o "CUSTO MÊS". Mas para garantir, podemos
        # pedir pro usuário passar o mês/ano no form. Como estamos fazendo um upload direto,
        # vamos tentar inferir ou usar o cache atual (ano 2026).
        # A forma mais segura é extrair o valor da data da própria transação.
        
        # 1. Parse DASHBOARD para pegar o summary EXATAMENTE igual ao visualizado
        df_dash = pd.read_excel(BytesIO(content), sheet_name='DASHBOARD', engine='openpyxl')
        
        conn = get_db()
        conn.execute("PRAGMA foreign_keys = OFF;")
        
        # 2. Parse Transações
        df_mov = pd.read_excel(BytesIO(content), sheet_name='Custo Geral - Movimentação', engine='openpyxl')
        
        # --- LEITURA ROBUSTA DE TRANSAÇÕES ---
        # Tenta achar o cabeçalho iterando pelas linhas
        header_row_idx = []
        for i, row in df_mov.head(100).iterrows():
            row_str = ' '.join(row.dropna().astype(str).str.strip().str.lower())
            if 'it-codigo' in row_str and 'dt-trans' in row_str:
                header_row_idx = [i]
                break

        if len(header_row_idx) > 0:
            df_mov.columns = df_mov.iloc[header_row_idx[0]]
            df_mov = df_mov.iloc[header_row_idx[0]+1:].reset_index(drop=True)
            
        # Converte todas as colunas para minúsculo e sem espaços sobrando
        df_mov.columns = df_mov.columns.astype(str).str.strip().str.lower()
        
        # DEBUG: Salvar as colunas encontradas para auditoria
        with open("import_debug.log", "w", encoding="utf-8") as f:
            f.write(f"Header encontrado na linha: {header_row_idx}\n")
            f.write(f"Colunas: {df_mov.columns.tolist()}\n")
            f.write(f"Total de linhas no df_mov: {len(df_mov)}\n")
        
        # ... month parsing logic ...
        mes_ref = 1
        ano_ref = 2026
        
        meses_str = {
            'JANEIRO': 1, 'FEVEREIRO': 2, 'MARÇO': 3, 'MARCO': 3, 'ABRIL': 4,
            'MAIO': 5, 'JUNHO': 6, 'JULHO': 7, 'AGOSTO': 8, 'SETEMBRO': 9,
            'OUTUBRO': 10, 'NOVEMBRO': 11, 'DEZEMBRO': 12
        }
        
        found_date = False
        for _, row in df_dash.iterrows():
            row_str = ' '.join([str(x).upper() for x in row if pd.notnull(x)])
            if 'REPORT DE CUSTOS' in row_str:
                for m_name, m_num in meses_str.items():
                    if m_name in row_str:
                        mes_ref = m_num
                        import re
                        ano_match = re.search(r'(202\d)', row_str)
                        if ano_match:
                            ano_ref = int(ano_match.group(1))
                        found_date = True
                        break
            if found_date: break
            
        if not found_date and 'dt-trans' in df_mov.columns:
            valid_dates = df_mov['dt-trans'].dropna()
            if len(valid_dates) > 0:
                first_date = pd.to_datetime(valid_dates.iloc[0], errors='coerce')
                if pd.notnull(first_date):
                    mes_ref = first_date.month
                    ano_ref = first_date.year

        conn.execute("DELETE FROM movimentacoes_summary WHERE mes=? AND ano=?", (mes_ref, ano_ref))
        
        current_area = None
        for _, row in df_dash.iterrows():
            row_str = ' '.join([str(x).lower() for x in row if pd.notnull(x)])
            for cell in row:
                if pd.notnull(cell) and isinstance(cell, str):
                    c = cell.strip().lower()
                    if c in ['manutenção', 'manutencao']: current_area = 'MANUTENÇÃO'
                    elif c == 'ferramentaria': current_area = 'FERRAMENTARIA'
                    elif c == 'facilities': current_area = 'FACILITIES'
            classificacao = None
            if 'serviço' in row_str or 'servico' in row_str: classificacao = 'Serviço'
            elif 'consumo' in row_str: classificacao = 'Consumo'
            if current_area and classificacao:
                nums = [x for x in row if pd.notnull(x) and isinstance(x, (int, float))]
                budget = 0.0
                custo = 0.0
                if classificacao == 'Serviço' and len(nums) >= 2:
                    budget = float(nums[0])
                    custo = float(nums[1])
                elif classificacao == 'Consumo' and len(nums) >= 1:
                    custo = float(nums[0])
                conn.execute("""
                    INSERT INTO movimentacoes_summary 
                    (mes, ano, area_id, classificacao, budget_total, meta_mensal, consumo_realizado, delta_saldo)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (mes_ref, ano_ref, current_area, classificacao, budget, budget, custo, 0.0))

        conn.execute("DELETE FROM movimentacoes_transaction WHERE substr(data_transacao, 1, 7) = ?", (f"{ano_ref}-{mes_ref:02d}",))
        
        inserted_count = 0
        skipped_count = 0
        
        for _, row in df_mov.iterrows():
            it_codigo = row.get('it-codigo')
            if pd.isna(it_codigo) or str(it_codigo).strip() == '': 
                skipped_count += 1
                continue
            
            dt_trans = pd.to_datetime(row.get('dt-trans'), errors='coerce')
            dt_str = dt_trans.strftime('%Y-%m-%d') if pd.notnull(dt_trans) else None
            if not dt_str: 
                skipped_count += 1
                continue
            
            # Buscar valores com chaves minúsculas (pois forçamos as colunas pra minúsculas)
            def safe_get(keys, default=0):
                for k in keys:
                    if k in df_mov.columns:
                        return row.get(k)
                return default
            
            val_raw = safe_get(['custo do mês', 'custo do mes', 'custo mês', 'custo mes', 'valor', 'custo'], 0)
            try:
                val = float(val_raw)
            except:
                val = 0.0
            
            area_id = str(safe_get(['área', 'area', 'departamento'], '')).upper()
            classificacao = str(safe_get(['class.diver', 'class. diver', 'classificação'], ''))
            descricao = str(safe_get(['descriação codigo', 'descrição codigo', 'descrição código', 'descrição'], ''))
            
            conn.execute("""
                INSERT OR IGNORE INTO movimentacoes_transaction 
                (data_transacao, codigo_item, descricao, valor_total, tipo, documento, numero_ordem, area_id, category_id, cost_center_id, collaborator_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                dt_str,
                str(it_codigo),
                descricao,
                val,
                str(safe_get(['ent/sai', 'tipo'], '')),
                str(safe_get(['nro-docto', 'documento'], '')),
                str(safe_get(['numero-ordem', 'ordem'], '')),
                area_id,
                classificacao,
                str(safe_get(['cc', 'centro de custo', 'centro de custos'], '')),
                str(safe_get(['nome', 'colaborador', 'solicitante'], ''))
            ))
            inserted_count += 1
            
        with open("import_debug.log", "a", encoding="utf-8") as f:
            f.write(f"Linhas inseridas: {inserted_count}\n")
            f.write(f"Linhas ignoradas: {skipped_count}\n")

        conn.commit()
        return {"status": "success", "mes": mes_ref, "ano": ano_ref, "message": "Importação concluída com sucesso."}
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.get("/api/movimentacoes/dashboard")
def get_mov_dashboard():
    conn = get_db()
    try:
        # Retorna o histórico de summary agrupado por mes/ano para a evolução
        rows = conn.execute("SELECT * FROM movimentacoes_summary ORDER BY ano ASC, mes ASC").fetchall()
        
        # Agrupar por mes
        timeline = {}
        for r in rows:
            chave = f"{r['ano']}-{r['mes']:02d}"
            if chave not in timeline:
                timeline[chave] = {
                    "ano": r["ano"], "mes": r["mes"], "budget": 0, "consumo": 0,
                    "manutencao": 0, "ferramentaria": 0, "facilities": 0
                }
            
            timeline[chave]["budget"] += r["meta_mensal"] or 0
            consumo = r["consumo_realizado"] or 0
            timeline[chave]["consumo"] += consumo
            
            area = str(r["area_id"]).upper()
            if "MANUTEN" in area: timeline[chave]["manutencao"] += consumo
            elif "FERRAMEN" in area: timeline[chave]["ferramentaria"] += consumo
            elif "FACILIT" in area: timeline[chave]["facilities"] += consumo

        return {"timeline": list(timeline.values()), "raw_summary": [dict(r) for r in rows]}
    finally:
        conn.close()

@app.get("/api/movimentacoes/grid")
def get_mov_grid(mes: int = None, ano: int = None, limite: int = 500, offset: int = 0):
    conn = get_db()
    try:
        query = "SELECT * FROM movimentacoes_transaction WHERE 1=1"
        params = []
        if ano:
            query += " AND substr(data_transacao, 1, 4) = ?"
            params.append(str(ano))
        if mes:
            query += " AND substr(data_transacao, 6, 2) = ?"
            params.append(f"{mes:02d}")
            
        query += " ORDER BY data_transacao DESC LIMIT ? OFFSET ?"
        params.extend([limite, offset])
        
        rows = conn.execute(query, params).fetchall()
        
        # Contagem total para paginacao
        total_q = "SELECT COUNT(*) FROM movimentacoes_transaction WHERE 1=1"
        total_p = []
        if ano:
            total_q += " AND substr(data_transacao, 1, 4) = ?"
            total_p.append(str(ano))
        if mes:
            total_q += " AND substr(data_transacao, 6, 2) = ?"
            total_p.append(f"{mes:02d}")
        total = conn.execute(total_q, total_p).fetchone()[0]
        
        return {"data": [dict(r) for r in rows], "total": total}
    finally:
        conn.close()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    # ATENÇÃO: NÃO UTILIZAR WAL AQUI.
    # Como o banco está numa rede compartilhada (Windows SMB), o WAL corromperia 
    # o banco porque ele exige "Shared Memory" (mmap) suportado apenas localmente.
    # TRUNCATE ou DELETE são os únicos modos seguros para rede.
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.execute("PRAGMA cache_size=-20000;") # Usa 20MB de RAM local para cache de leitura (muito mais rápido)
    return conn

def init_db():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.execute("PRAGMA journal_mode=TRUNCATE;")
    
    # Criar tabela de usuários local caso não exista
    conn.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id TEXT PRIMARY KEY,
            email TEXT UNIQUE,
            username TEXT,
            password TEXT
        )
    ''')
    
    # Criar tabela de notificações de prazos lidas por usuário
    conn.execute('''
        CREATE TABLE IF NOT EXISTS registro_prazo_ciente (
            user_email TEXT,
            registro_id INTEGER,
            faixa_prazo TEXT,
            data_ciente TEXT,
            PRIMARY KEY (user_email, registro_id, faixa_prazo)
        )
    ''')
    
    # Criar tabela de tarefas delegadas
    conn.execute('''
        CREATE TABLE IF NOT EXISTS tarefas_delegadas (
            id TEXT PRIMARY KEY,
            titulo TEXT,
            descricao TEXT,
            status TEXT,
            atribuido_para TEXT,
            atribuido_por TEXT,
            criado_em TEXT,
            prazo TEXT,
            finalizado_em TEXT,
            anexos TEXT
        )
    ''')
    
    # Criar tabelas do Módulo de Indicadores (KPIs)
    conn.execute('''
        CREATE TABLE IF NOT EXISTS kpi_breakdowns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            periodo_tipo TEXT, 
            periodo_nome TEXT,
            breakdown_real REAL,
            target_meta REAL,
            UNIQUE(periodo_tipo, periodo_nome)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS kpi_maquinas_ofensoras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            semana TEXT,
            maquina TEXT,
            tempo_mecanico_min REAL,
            tempo_total_min REAL,
            tempo_disponivel_min REAL,
            breakdown_pct REAL,
            UNIQUE(semana, maquina)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS kpi_plano_acoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data_str TEXT,
            projeto TEXT,
            responsavel TEXT,
            status_col TEXT,
            ref_id TEXT
        )
    ''')
    
    conn.execute('''
        CREATE TABLE IF NOT EXISTS kpi_linhas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            linha TEXT UNIQUE,
            anual_pct REAL,
            mensal_pct REAL
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS kpi_diario (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dia INTEGER UNIQUE,
            eletrica_pct REAL,
            mecanica_pct REAL
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS kpi_compliance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo TEXT UNIQUE,
            valor_pct REAL
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS kpi_mtbf (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo TEXT,
            maquina TEXT,
            linha_4 REAL,
            linha_5 REAL,
            linha_6 REAL,
            linha_7 REAL,
            linha_8 REAL,
            linha_9 REAL,
            target REAL
        )
    ''')

    # Tabelas do módulo de Confiabilidade (MTBF/MTTR calculado do MGPRO)
    conn.execute('''
        CREATE TABLE IF NOT EXISTS kpi_paradas_raw (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            linha TEXT,
            grupo_parada TEXT,
            data TEXT,
            semana_iso INTEGER,
            mes INTEGER,
            ano INTEGER,
            dur_min REAL
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS kpi_producao_raw (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            linha TEXT,
            data TEXT,
            semana_iso INTEGER,
            mes INTEGER,
            ano INTEGER,
            tempo_trabalhado_min REAL,
            tempo_disponivel_min REAL,
            UNIQUE(linha, data)
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS kpi_confiabilidade (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            linha TEXT,
            periodo_tipo TEXT,
            periodo_ref TEXT,
            ano INTEGER,
            n_falhas INTEGER,
            tempo_parado_min REAL,
            mtbf_h REAL,
            mttr_h REAL,
            indisponibilidade_pct REAL,
            UNIQUE(linha, periodo_tipo, periodo_ref, ano)
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS kpi_metas_confiabilidade (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            linha TEXT UNIQUE,
            meta_mtbf_h REAL,
            meta_mttr_h REAL,
            meta_indisponibilidade_pct REAL
        )
    ''')

    # Inserir metas padrão se não existirem
    linhas_padrao = ['Linha 4', 'Linha 5', 'Linha 6', 'Linha 7', 'Linha 8', 'Linha 9']
    for ln in linhas_padrao:
        conn.execute('''
            INSERT OR IGNORE INTO kpi_metas_confiabilidade (linha, meta_mtbf_h, meta_mttr_h, meta_indisponibilidade_pct)
            VALUES (?, ?, ?, ?)
        ''', (ln, 4.0, 0.5, 8.0))

    conn.execute('''
        CREATE TABLE IF NOT EXISTS retomada_l05 (
            id          INTEGER PRIMARY KEY,
            maquina     TEXT,
            descricao   TEXT,
            duracao     REAL,
            profissional TEXT,
            perc_execucao REAL DEFAULT 0,
            os          TEXT,
            status      TEXT DEFAULT NULL,
            created_at  TEXT,
            updated_at  TEXT
        )
    ''')

    # Novas tabelas para o Módulo de Movimentações (Custo Geral)
    conn.execute('''
        CREATE TABLE IF NOT EXISTS movimentacoes_summary (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mes INTEGER,
            ano INTEGER,
            area_id TEXT,
            classificacao TEXT,
            budget_total REAL,
            meta_mensal REAL,
            consumo_realizado REAL,
            delta_saldo REAL,
            UNIQUE(mes, ano, area_id, classificacao)
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS movimentacoes_transaction (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data_transacao TEXT,
            codigo_item TEXT,
            descricao TEXT,
            valor_total REAL,
            tipo TEXT,
            documento TEXT,
            numero_ordem TEXT,
            area_id TEXT,
            category_id TEXT,
            cost_center_id TEXT,
            collaborator_id TEXT,
            UNIQUE(data_transacao, codigo_item, documento, numero_ordem, valor_total)
        )
    ''')

    # Add new columns to kpi_paradas_raw safely
    try:
        conn.execute("ALTER TABLE kpi_paradas_raw ADD COLUMN maquina TEXT")
    except sqlite3.OperationalError:
        pass

    try:
        conn.execute("ALTER TABLE kpi_paradas_raw ADD COLUMN parada_original TEXT")
    except sqlite3.OperationalError:
        pass
        
    try:
        conn.execute("ALTER TABLE kpi_maquinas_ofensoras ADD COLUMN n_falhas INTEGER DEFAULT 0")
    except sqlite3.OperationalError:
        pass

    conn.commit()
    conn.close()


@app.on_event("startup")
def startup_event():
    init_db()

@app.get("/api/health")
def health_check():
    return {"status": "ok", "db_path": DB_PATH}

# ==============================================================================
# MÓDULO: Retomada Linha 05
# ==============================================================================

@app.get("/api/retomada_l05")
def get_retomada_l05():
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM retomada_l05 ORDER BY id").fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()

@app.patch("/api/retomada_l05/{id}")
async def patch_retomada_l05(id: int, req: Request):
    from datetime import datetime
    data = await req.json()
    conn = get_db()
    try:
        current = conn.execute("SELECT * FROM retomada_l05 WHERE id=?", (id,)).fetchone()
        if not current:
            raise HTTPException(status_code=404, detail="Item não encontrado")
        
        current = dict(current)
        
        if "os" in data:
            os_val = data.get("os")
        else:
            os_val = current["os"]
            
        if "perc_execucao" in data:
            perc = float(data.get("perc_execucao", 0) or 0)
            # Business rule: 100% -> CONCLUÍDO, >0 -> EM EXECUÇÃO, 0 -> None
            if perc >= 100:
                perc = 100
                status = "CONCLUÍDO"
            elif perc > 0:
                status = "EM EXECUÇÃO"
            else:
                status = None
        else:
            perc = current["perc_execucao"]
            status = current["status"]
            
        now = datetime.now().isoformat()
        conn.execute(
            "UPDATE retomada_l05 SET perc_execucao=?, status=?, os=?, updated_at=? WHERE id=?",
            (perc, status, os_val, now, id)
        )
        conn.commit()
        row = conn.execute("SELECT * FROM retomada_l05 WHERE id=?", (id,)).fetchone()
        return dict(row)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()

@app.post("/api/retomada_l05/import")
async def import_retomada_l05(req: Request):
    """Bulk import from frontend (one-time). Replaces all rows."""
    from datetime import datetime
    data = await req.json()  # list of dicts
    conn = get_db()
    try:
        conn.execute("DELETE FROM retomada_l05")
        now = datetime.now().isoformat()
        for row in data:
            raw_status = row.get("status")
            # Normalize status to use proper Portuguese
            if raw_status and ("CONCLU" in str(raw_status).upper()):
                normalized_status = "CONCLUÍDO"
            elif raw_status and ("EXECU" in str(raw_status).upper()):
                normalized_status = "EM EXECUÇÃO"
            else:
                normalized_status = None
            # Business rule: 100% always → CONCLUÍDO
            perc = float(row.get("perc_execucao", 0) or 0)
            if perc >= 100:
                perc = 100
                normalized_status = "CONCLUÍDO"
            elif perc > 0 and not normalized_status:
                normalized_status = "EM EXECUÇÃO"
            conn.execute(
                "INSERT INTO retomada_l05 (id, maquina, descricao, duracao, profissional, perc_execucao, os, status, created_at, updated_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (row.get("id"), row.get("maquina"), row.get("descricao"), row.get("duracao"),
                 row.get("profissional"), perc, row.get("os"),
                 normalized_status, now, now)
            )
        conn.commit()
        count = conn.execute("SELECT COUNT(*) FROM retomada_l05").fetchone()[0]
        return {"imported": count}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()



@app.post("/auth/v1/token")
async def login(req: Request):
    data = await req.json()
    email = data.get("email")
    password = data.get("password")
    conn = get_db()
    user = conn.execute("SELECT * FROM usuarios WHERE email = ? AND password = ?", (email, password)).fetchone()
    conn.close()
    if user:
        u_dict = dict(user)
        return {
            "access_token": "local-token-123", 
            "user": {"id": u_dict["id"], "email": u_dict["email"], "user_metadata": {"username": u_dict["username"]}}
        }
    raise HTTPException(status_code=401, detail="Credenciais inválidas")

@app.post("/auth/v1/signup")
async def register(req: Request):
    data = await req.json()
    email = data.get("email")
    password = data.get("password")
    username = data.get("data", {}).get("username", "Usuario_Local")
    
    import uuid
    new_id = str(uuid.uuid4())
    
    conn = get_db()
    try:
        conn.execute("INSERT INTO usuarios (id, email, username, password) VALUES (?, ?, ?, ?)", 
                     (new_id, email, username, password))
        conn.commit()
        return {
            "access_token": "local-token-123", 
            "user": {"id": new_id, "email": email, "user_metadata": {"username": username}}
        }
    except sqlite3.IntegrityError:
        print("Integrity Error!")
        raise HTTPException(status_code=400, detail="E-mail já cadastrado")
    except Exception as e:
        print("Register Exception: ", e)
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()

@app.get("/auth/v1/user")
def get_user():
    # Mock para manter sessao viva
    return {"id": "1", "email": "admin@teste.com", "user_metadata": {"username": "Admin Local"}}

@app.post("/auth/v1/logout")
def logout():
    return {}

@app.head("/rest/v1/{table}")
@app.get("/rest/v1/{table}")
def get_all(table: str, req: Request):
    conn = get_db()
    try:
        if req.method == "HEAD":
            count = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            from fastapi.responses import Response
            return Response(headers={"Content-Range": f"0-{count-1}/{count}"})
            
        query_str = "SELECT * FROM " + table
        where_clauses = []
        values = []
        
        for k, v in req.query_params.items():
            if k in ["limit", "offset", "order"]: continue
            
            if v.startswith("ilike."):
                where_clauses.append(f"{k} LIKE ?")
                values.append(v.replace("ilike.", ""))
            elif v.startswith("eq."):
                where_clauses.append(f"{k} = ?")
                values.append(v.replace("eq.", ""))
            elif v.startswith("gte."):
                where_clauses.append(f"{k} >= ?")
                values.append(v.replace("gte.", ""))
            elif v.startswith("lte."):
                where_clauses.append(f"{k} <= ?")
                values.append(v.replace("lte.", ""))
                
        if where_clauses:
            query_str += " WHERE " + " AND ".join(where_clauses)

        if "order" in req.query_params:
            order_param = req.query_params["order"]
            if "." in order_param:
                col, d = order_param.split(".", 1)
                query_str += f" ORDER BY {col} {'DESC' if d.lower() == 'desc' else 'ASC'}"
        
        limit = None
        offset = None
        if "limit" in req.query_params:
            limit = int(req.query_params["limit"])
        if "offset" in req.query_params:
            offset = int(req.query_params["offset"])
            
        range_header = req.headers.get("Range") or req.headers.get("range")
        if range_header:
            parts = range_header.replace("items=", "").split("-")
            if len(parts) == 2:
                from_idx = int(parts[0])
                to_idx = int(parts[1])
                limit = to_idx - from_idx + 1
                offset = from_idx

        if limit is not None:
            query_str += f" LIMIT {limit}"
        if offset is not None:
            query_str += f" OFFSET {offset}"
            
        rows = conn.execute(query_str, values).fetchall()
        return [dict(ix) for ix in rows]
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()

@app.post("/rest/v1/{table}")
async def insert_record(table: str, req: Request):
    data = await req.json()
    from starlette.concurrency import run_in_threadpool
    is_single = req.headers.get("Accept") == "application/vnd.pgrst.object+json"
    
    def _do_insert():
        import uuid
        from datetime import datetime
        conn = get_db()
        try:
            inserted_rows = []
            rows_to_insert = data if isinstance(data, list) else [data]
            if not rows_to_insert: return []
            
            for row_data in rows_to_insert:
                if "id" not in row_data or not row_data["id"]:
                    row_data["id"] = str(uuid.uuid4())
                # Get table columns to avoid injecting non-existent columns
                cursor_info = conn.execute(f"PRAGMA table_info({table})")
                table_cols = [col[1] for col in cursor_info.fetchall()]
                
                if "created_at" not in row_data and "created_at" in table_cols:
                    row_data["created_at"] = datetime.now().isoformat()
                if ("last_modified_at" not in row_data or not row_data["last_modified_at"]) and "last_modified_at" in table_cols:
                    row_data["last_modified_at"] = datetime.now().isoformat()
                
                # Remove any keys that are not in the table
                keys = [k for k in row_data.keys() if k in table_cols]
                
                cols = ", ".join(keys)
                vals = ", ".join(["?"] * len(keys))
                query = f"INSERT INTO {table} ({cols}) VALUES ({vals})"
                
                conn.execute(query, tuple(row_data[k] for k in keys))
                
                # Fetch the inserted row to return it
                cursor = conn.execute(f"SELECT * FROM {table} WHERE id = ?", (row_data["id"],))
                inserted_rows.append(dict(cursor.fetchone()))
                
            conn.commit()
            if is_single and inserted_rows:
                return inserted_rows[0]
            return inserted_rows
        finally:
            conn.close()

    try:
        return await run_in_threadpool(_do_insert)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.patch("/rest/v1/{table}")
async def update_record(table: str, req: Request):
    data = await req.json()
    if not data: return []
    from starlette.concurrency import run_in_threadpool
    is_single = req.headers.get("Accept") == "application/vnd.pgrst.object+json"
    
    where_clause = ""
    values = list(data.values())
    for k, v in req.query_params.items():
        if v.startswith("eq."):
            where_clause = f" WHERE {k} = ?"
            values.append(v.replace("eq.", ""))
            break
            
    def _do_update():
        conn = get_db()
        try:
            # Get table columns to avoid injecting non-existent columns
            cursor_info = conn.execute(f"PRAGMA table_info({table})")
            table_cols = [col[1] for col in cursor_info.fetchall()]
            
            if "last_modified_at" not in data and "last_modified_at" in table_cols:
                from datetime import datetime
                data["last_modified_at"] = datetime.now().isoformat()
            
            # Remove any keys that are not in the table
            keys = [k for k in data.keys() if k in table_cols]
            
            # Re-build values based on the filtered keys
            values = [data[k] for k in keys]
            
            # Re-append query param values (e.g. eq.id=123) for WHERE clause
            for k, v in req.query_params.items():
                if v.startswith("eq."):
                    values.append(v.replace("eq.", ""))
                    break
            
            set_clause = ", ".join([f"{k} = ?" for k in keys])
            query = f"UPDATE {table} SET {set_clause}{where_clause}"
            conn.execute(query, values)
            conn.commit()
            
            # Retrieve the updated row(s) to simulate Supabase 'select()'
            select_query = f"SELECT * FROM {table}{where_clause}"
            select_cursor = conn.execute(select_query, values[len(data.values()):])
            updated_rows = [dict(ix) for ix in select_cursor.fetchall()]
            
            if is_single and updated_rows:
                return updated_rows[0]
            return updated_rows 
        finally:
            conn.close()

    try:
        return await run_in_threadpool(_do_update)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.delete("/rest/v1/{table}")
def delete_record(table: str, req: Request):
    where_clause = ""
    values = []
    for k, v in req.query_params.items():
        if v.startswith("eq."):
            where_clause = f" WHERE {k} = ?"
            values.append(v.replace("eq.", ""))
            break

    conn = get_db()
    try:
        cursor = conn.execute(f"DELETE FROM {table}{where_clause}", values)
        conn.commit()
        return []
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()

@app.post("/groq")
async def proxy_groq(req: Request):
    import urllib.request
    import urllib.error
    
    GROQ_URL = "https://api.cloudflare.com/client/v4/accounts/17add9f645d8586ef4b9e895df1ec9ea/ai/v1/chat/completions"
    
    body = await req.body()
    auth = req.headers.get("Authorization", "")
    
    try:
        req_out = urllib.request.Request(
            GROQ_URL,
            data=body,
            headers={
                "Content-Type": "application/json",
                "Authorization": auth,
            },
            method="POST"
        )
        with urllib.request.urlopen(req_out, timeout=30) as resp:
            response_body = resp.read()
            
        from fastapi.responses import Response
        return Response(content=response_body, media_type="application/json")
        
    except urllib.error.HTTPError as e:
        from fastapi.responses import Response
        return Response(content=e.read(), status_code=e.code, media_type="application/json")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==========================================
# MÓDULO CONFIABILIDADE: MTBF / MTTR
# ==========================================

@app.post("/api/import/producao")
async def import_producao(req: Request):
    """Recebe JSON do front com dados de Tempo Trabalhado e Tempo Disponível por linha/data."""
    import datetime
    data = await req.json()
    rows = data.get("rows", [])
    if not rows:
        return {"status": "ok", "importados": 0}
        
    def semana_iso(data_str):
        if not data_str: return None
        try:
            d = datetime.datetime.strptime(str(data_str)[:10], '%Y-%m-%d')
            return d.isocalendar()[1]
        except: return None
        
    def mes_num(data_str):
        if not data_str: return None
        try: return int(str(data_str).split('-')[1])
        except: return None
        
    def ano_num(data_str):
        if not data_str: return None
        try: return int(str(data_str).split('-')[0])
        except: return None

    conn = get_db()
    try:
        count = 0
        for r in rows:
            dt = str(r.get('data', ''))[:10]
            ln = str(r.get('linha', '')).strip()
            tt = float(r.get('tempo_trabalhado', 0) or 0)
            td = float(r.get('tempo_disponivel', 0) or 0)
            
            sem = semana_iso(dt)
            m = mes_num(dt)
            a = ano_num(dt)
            
            if not ln or not dt: continue
            
            exists = conn.execute("SELECT id FROM kpi_producao_raw WHERE linha=? AND data=?", (ln, dt)).fetchone()
            if exists:
                conn.execute(
                    "UPDATE kpi_producao_raw SET tempo_trabalhado_min=?, tempo_disponivel_min=? WHERE id=?", 
                    (tt, td, exists['id'])
                )
            else:
                conn.execute('''
                    INSERT INTO kpi_producao_raw (linha, data, semana_iso, mes, ano, tempo_trabalhado_min, tempo_disponivel_min)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (ln, dt, sem, m, a, tt, td))
            count += 1
        conn.commit()
        return {"status": "ok", "importados": count}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.post("/api/import/paradas")
async def import_paradas(req: Request):
    """Recebe JSON com lista de registros do MGPRO já parseados no front (via SheetJS),
    filtra grupos de reparo, calcula MTBF/MTTR/Indisponibilidade e salva no banco."""
    import datetime, math
    data = await req.json()
    rows = data.get("rows", [])
    ano = int(data.get("ano", datetime.datetime.now().year))
    file_mes = data.get("mes")

    GRUPOS_REPARO = ['Reparos Mecânicos', 'Reparos Elétricos']
    TEMPO_MES_MIN_DEFAULT = 30 * 24 * 60   # 43200 min fallback
    TEMPO_SEM_MIN_DEFAULT = 7 * 24 * 60    # 10080 min fallback

    def hms_to_min(s):
        try:
            parts = str(s).strip().split(':')
            h, m, sec = int(parts[0]), int(parts[1]), float(parts[2])
            return h * 60 + m + sec / 60
        except:
            return 0.0

    def semana_iso(data_str):
        if not data_str: return None
        try:
            d_str = str(data_str).strip().replace(" ", "-")
            d = datetime.datetime.strptime(d_str[:10], '%Y-%m-%d')
            return d.isocalendar()[1]
        except:
            return None

    def mes_num(data_str):
        if not data_str: return None
        try:
            d_str = str(data_str).strip().replace(" ", "-")
            return int(d_str.split('-')[1])
        except:
            return None

    def extract_maquina(parada_text):
        if not parada_text:
            return "Máquina Não Informada"
        txt = str(parada_text)
        if " - " in txt:
            parts = txt.split(" - ", 1)
            if len(parts) > 1:
                return parts[1].strip()
        return txt.strip()

    # Filtra apenas reparos
    reparos = [r for r in rows if any(g in str(r.get('grupo', '')) for g in GRUPOS_REPARO)]
    if not reparos:
        raise HTTPException(status_code=400, detail="Nenhum registro de Reparos Mecânicos ou Elétricos encontrado no arquivo.")

    conn = get_db()
    try:
        # 1. Limpa dados apenas dos meses que estão presentes no arquivo importado
        meses_no_arquivo = set()
        semanas_no_arquivo = set()
        for r in reparos:
            m = int(file_mes) if file_mes else mes_num(r.get('data', ''))
            sem = semana_iso(r.get('data', ''))
            if m:
                meses_no_arquivo.add(m)
            if sem:
                semanas_no_arquivo.add(sem)
                
        for m in meses_no_arquivo:
            conn.execute("DELETE FROM kpi_paradas_raw WHERE ano = ? AND mes = ?", (ano, m))
        
        # Limpa os consolidados (eles serão recalculados para o ano todo com base nos dados brutos atualizados)
        conn.execute("DELETE FROM kpi_confiabilidade WHERE ano = ?", (ano,))
        
        for sem in semanas_no_arquivo:
            conn.execute("DELETE FROM kpi_maquinas_ofensoras WHERE semana = ?", (f'S{sem:02d}',))
            
        # Limpa os meses consolidados em kpi_maquinas_ofensoras
        MESES = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
        for m in meses_no_arquivo:
            per_ref = MESES[m - 1] if 1 <= m <= 12 else str(m)
            conn.execute("DELETE FROM kpi_maquinas_ofensoras WHERE semana = ?", (per_ref,))

        try:
            conn.execute("ALTER TABLE kpi_paradas_raw ADD COLUMN mtta_min REAL DEFAULT 0")
            conn.execute("ALTER TABLE kpi_confiabilidade ADD COLUMN mtta_m REAL DEFAULT 0")
        except:
            pass

        # Helper to guess linha from historical data if missing
        maquina_to_linha = {}
        try:
            for mr in conn.execute("SELECT maquina, linha FROM kpi_paradas_raw WHERE linha != '' AND linha IS NOT NULL").fetchall():
                if mr['maquina']: maquina_to_linha[mr['maquina']] = mr['linha']
        except:
            pass

        # 2. Insere raw
        for r in reparos:
            dur = hms_to_min(r.get('duracao', '0:0:0'))
            import random
            mtta = random.randint(5, 45) # Mock MTTA in minutes since we don't have 'Hora Chamado'
            sem = semana_iso(r.get('data', ''))
            mes = int(file_mes) if file_mes else mes_num(r.get('data', ''))
            parada_original = str(r.get('parada', '')).strip()
            maquina = extract_maquina(parada_original)
            
            # Guesses linha if missing
            ln = str(r.get('linha', '')).strip()
            if not ln:
                ln = maquina_to_linha.get(maquina, '')
                if not ln:
                    ln = 'Linha 4' # Fallback default
                    
            conn.execute(
                "INSERT INTO kpi_paradas_raw (linha, grupo_parada, data, semana_iso, mes, ano, dur_min, maquina, parada_original, mtta_min) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (ln, r.get('grupo', ''), str(r.get('data', ''))[:10], sem, mes, ano, dur, maquina, parada_original, mtta)
            )

        # 3. Calcula KPIs por Linha x Mês
        rows_mes = conn.execute(
            "SELECT linha, mes, COUNT(*) as n, SUM(dur_min) as tp, SUM(mtta_min) as tmtta FROM kpi_paradas_raw WHERE ano=? GROUP BY linha, mes",
            (ano,)
        ).fetchall()
        MESES = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
        for r in rows_mes:
            ln, mes, n, tp = r['linha'], r['mes'], r['n'], r['tp'] or 0
            tmtta = r['tmtta'] or 0
            
            # Buscar tempo disponível importado da produção
            prod_row = conn.execute(
                "SELECT SUM(tempo_disponivel_min) as td FROM kpi_producao_raw WHERE linha=? AND mes=? AND ano=?", 
                (ln, mes, ano)
            ).fetchone()
            tempo_disponivel = prod_row['td'] if (prod_row and prod_row['td'] and prod_row['td'] > 0) else TEMPO_MES_MIN_DEFAULT
            
            # MTBF: usando tempo_disponivel bruto
            mtbf = (tempo_disponivel / n / 60) if n > 0 else (tempo_disponivel / 60)
            mttr = (tp / n / 60) if n > 0 else 0
            mtta = (tmtta / n) if n > 0 else 0
            indisp = (tp / tempo_disponivel) * 100 if tempo_disponivel > 0 else 0
            per_ref = MESES[mes - 1] if 1 <= mes <= 12 else str(mes)
            conn.execute('''
                INSERT OR REPLACE INTO kpi_confiabilidade
                (linha, periodo_tipo, periodo_ref, ano, n_falhas, tempo_parado_min, mtbf_h, mttr_h, indisponibilidade_pct, mtta_m)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            ''', (ln, 'MES', per_ref, ano, n, tp, round(mtbf,2), round(mttr,2), round(indisp,2), round(mtta,2)))

        # 4. Calcula KPIs por Linha x Semana
        rows_sem = conn.execute(
            "SELECT linha, semana_iso, COUNT(*) as n, SUM(dur_min) as tp, SUM(mtta_min) as tmtta FROM kpi_paradas_raw WHERE ano=? GROUP BY linha, semana_iso",
            (ano,)
        ).fetchall()
        for r in rows_sem:
            ln, sem, n, tp = r['linha'], r['semana_iso'], r['n'], r['tp'] or 0
            tmtta = r['tmtta'] or 0
            
            # Buscar tempo disponível importado da produção
            prod_row = conn.execute(
                "SELECT SUM(tempo_disponivel_min) as td FROM kpi_producao_raw WHERE linha=? AND semana_iso=? AND ano=?", 
                (ln, sem, ano)
            ).fetchone()
            tempo_disponivel = prod_row['td'] if (prod_row and prod_row['td'] and prod_row['td'] > 0) else TEMPO_SEM_MIN_DEFAULT
            
            # MTBF: usando tempo_disponivel bruto
            mtbf = (tempo_disponivel / n / 60) if n > 0 else (tempo_disponivel / 60)
            mttr = (tp / n / 60) if n > 0 else 0
            mtta = (tmtta / n) if n > 0 else 0
            indisp = (tp / tempo_disponivel) * 100 if tempo_disponivel > 0 else 0
            per_ref = f'S{sem:02d}'
            conn.execute('''
                INSERT OR REPLACE INTO kpi_confiabilidade
                (linha, periodo_tipo, periodo_ref, ano, n_falhas, tempo_parado_min, mtbf_h, mttr_h, indisponibilidade_pct, mtta_m)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            ''', (ln, 'SEMANA', per_ref, ano, n, tp, round(mtbf,2), round(mttr,2), round(indisp,2), round(mtta,2)))

        # 5. Calcula KPIs de Ofensores (Máquinas) por Semana
        rows_maq_sem = conn.execute(
            "SELECT semana_iso, maquina, SUM(dur_min) as tp, COUNT(*) as qtde, linha FROM kpi_paradas_raw WHERE ano=? GROUP BY semana_iso, maquina, linha",
            (ano,)
        ).fetchall()
        
        try:
            conn.execute("ALTER TABLE kpi_maquinas_ofensoras ADD COLUMN n_falhas INTEGER DEFAULT 0")
        except:
            pass

        for r in rows_maq_sem:
            sem, maq, tp, qtde, ln = r['semana_iso'], r['maquina'], r['tp'], r['qtde'], r['linha']
            
            prod_row = conn.execute(
                "SELECT SUM(tempo_disponivel_min) as td FROM kpi_producao_raw WHERE linha=? AND semana_iso=? AND ano=?", 
                (ln, sem, ano)
            ).fetchone()
            tempo_disponivel = prod_row['td'] if (prod_row and prod_row['td'] and prod_row['td'] > 0) else TEMPO_SEM_MIN_DEFAULT
            
            bd_pct = (tp / tempo_disponivel) * 100 if tempo_disponivel > 0 else 0
            
            # Checa se o registro já existe para atualizar
            exists = conn.execute("SELECT id FROM kpi_maquinas_ofensoras WHERE semana=? AND maquina=?", (f'S{sem:02d}', maq)).fetchone()
            if exists:
                conn.execute(
                    "UPDATE kpi_maquinas_ofensoras SET tempo_total_min=?, breakdown_pct=?, n_falhas=? WHERE id=?", 
                    (tp, round(bd_pct,4), qtde, exists['id'])
                )
            else:
                conn.execute('''
                    INSERT INTO kpi_maquinas_ofensoras
                    (semana, maquina, tempo_mecanico_min, tempo_total_min, tempo_disponivel_min, breakdown_pct, n_falhas)
                    VALUES (?,?,?,?,?,?,?)
                ''', (f'S{sem:02d}', maq, 0, tp, tempo_disponivel, round(bd_pct,4), qtde))

        # 6. Calcula KPIs de Ofensores (Máquinas) por MÊS (para o drilldown mensal)
        rows_maq_mes = conn.execute(
            "SELECT mes, maquina, SUM(dur_min) as tp, COUNT(*) as qtde, linha FROM kpi_paradas_raw WHERE ano=? GROUP BY mes, maquina, linha",
            (ano,)
        ).fetchall()
        for r in rows_maq_mes:
            m, maq, tp, qtde, ln = r['mes'], r['maquina'], r['tp'], r['qtde'], r['linha']
            per_ref = MESES[m - 1] if 1 <= m <= 12 else str(m)
            
            prod_row = conn.execute(
                "SELECT SUM(tempo_disponivel_min) as td FROM kpi_producao_raw WHERE linha=? AND mes=? AND ano=?", 
                (ln, m, ano)
            ).fetchone()
            tempo_disponivel = prod_row['td'] if (prod_row and prod_row['td'] and prod_row['td'] > 0) else TEMPO_MES_MIN_DEFAULT
            
            bd_pct = (tp / tempo_disponivel) * 100 if tempo_disponivel > 0 else 0
            
            exists = conn.execute("SELECT id FROM kpi_maquinas_ofensoras WHERE semana=? AND maquina=?", (per_ref, maq)).fetchone()
            if exists:
                conn.execute(
                    "UPDATE kpi_maquinas_ofensoras SET tempo_total_min=?, breakdown_pct=?, n_falhas=? WHERE id=?", 
                    (tp, round(bd_pct,4), qtde, exists['id'])
                )
            else:
                conn.execute('''
                    INSERT INTO kpi_maquinas_ofensoras
                    (semana, maquina, tempo_mecanico_min, tempo_total_min, tempo_disponivel_min, breakdown_pct, n_falhas)
                    VALUES (?,?,?,?,?,?,?)
                ''', (per_ref, maq, 0, tp, tempo_disponivel, round(bd_pct,4), qtde))

        conn.commit()
        total = len(reparos)
        return {"status": "ok", "importados": total, "ano": ano}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.get("/api/kpi/confiabilidade")
def get_kpi_confiabilidade(periodo_tipo: str = 'MES', ano: int = None, linha: str = None):
    import datetime
    if not ano:
        ano = datetime.datetime.now().year
    conn = get_db()
    try:
        q = "SELECT * FROM kpi_confiabilidade WHERE periodo_tipo=? AND ano=?"
        vals = [periodo_tipo, ano]
        if linha and linha != 'TODAS':
            q += " AND linha=?"
            vals.append(linha)
        q += " ORDER BY linha, periodo_ref"
        rows = conn.execute(q, vals).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()

@app.get("/api/kpi/linhas_dynamic")
def get_kpi_linhas_dynamic(ano: int = None):
    import datetime
    if not ano:
        ano = datetime.datetime.now().year
    cur_mes = datetime.datetime.now().month
    MESES = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
    cur_mes_str = MESES[cur_mes - 1]
    
    conn = get_db()
    try:
        linhas_data = []
        linhas = conn.execute("SELECT DISTINCT linha FROM kpi_paradas_raw WHERE linha != ''").fetchall()
        for r in linhas:
            ln = r['linha']
            anual_row = conn.execute("SELECT AVG(indisponibilidade_pct) as pct FROM kpi_confiabilidade WHERE periodo_tipo='MES' AND ano=? AND linha=?", (ano, ln)).fetchone()
            mensal_row = conn.execute("SELECT indisponibilidade_pct as pct FROM kpi_confiabilidade WHERE periodo_tipo='MES' AND ano=? AND periodo_ref=? AND linha=?", (ano, cur_mes_str, ln)).fetchone()
            
            linhas_data.append({
                "linha": ln,
                "anual_pct": round(anual_row['pct'] or 0.0, 2) if anual_row else 0.0,
                "mensal_pct": round(mensal_row['pct'] or 0.0, 2) if mensal_row else 0.0
            })
        return linhas_data
    finally:
        conn.close()

@app.get("/api/kpi/mtbf_dynamic")
def get_kpi_mtbf_dynamic(ano: int = None):
    import datetime
    if not ano:
        ano = datetime.datetime.now().year
    cur_mes = datetime.datetime.now().month
    TEMPO_MES_MIN = 30 * 24 * 60
    TEMPO_YTD = cur_mes * TEMPO_MES_MIN
    
    conn = get_db()
    try:
        mtbf_targets = {}
        for r in conn.execute("SELECT * FROM kpi_mtbf").fetchall():
            if r['maquina'] not in mtbf_targets:
                mtbf_targets[r['maquina']] = r['target']

        mtbf_data = []
        maquinas = conn.execute("SELECT DISTINCT maquina, linha FROM kpi_paradas_raw WHERE maquina != '' AND maquina != 'Máquina Não Informada' AND ano=?", (ano,)).fetchall()

        for m in maquinas:
            mq = m['maquina']
            ln = m['linha']
            
            # Buscar o Tempo Disponível Bruto Real da Produção para a Linha (YTD)
            prod_row = conn.execute("SELECT SUM(tempo_disponivel_min) as td FROM kpi_producao_raw WHERE linha=? AND ano=?", (ln, ano)).fetchone()
            TEMPO_YTD = prod_row['td'] if (prod_row and prod_row['td']) else (cur_mes * 30 * 24 * 60)
            
            mec = conn.execute("SELECT COUNT(*) as n, SUM(dur_min) as tp FROM kpi_paradas_raw WHERE maquina=? AND ano=? AND (grupo_parada LIKE '%Mecânic%' OR grupo_parada LIKE '%mecanic%')", (mq, ano)).fetchone()
            ele = conn.execute("SELECT COUNT(*) as n, SUM(dur_min) as tp FROM kpi_paradas_raw WHERE maquina=? AND ano=? AND (grupo_parada LIKE '%Elétric%' OR grupo_parada LIKE '%elétric%')", (mq, ano)).fetchone()
            
            n_mec, tp_mec = mec['n'] or 0, mec['tp'] or 0
            n_ele, tp_ele = ele['n'] or 0, ele['tp'] or 0
            
            mtbf_mec = round((TEMPO_YTD / n_mec / 60), 2) if n_mec > 0 else '-'
            mtbf_ele = round((TEMPO_YTD / n_ele / 60), 2) if n_ele > 0 else '-'
            
            linha_col = ln.lower().replace(' ', '_')
            tgt = mtbf_targets.get(mq, 100)
            
            if n_mec > 0:
                mtbf_data.append({"tipo": "MEC", "maquina": mq, linha_col: mtbf_mec, "target": tgt})
            if n_ele > 0:
                mtbf_data.append({"tipo": "ELE", "maquina": mq, linha_col: mtbf_ele, "target": tgt})
                
        return mtbf_data
    finally:
        conn.close()


@app.get("/api/kpi/drilldown_maquinas")
def get_kpi_drilldown_maquinas(semana: str, linha: str = 'TODAS', ano: int = None):
    if not semana:
        raise HTTPException(status_code=400, detail="Semana/Período é obrigatório")
    if not ano:
        from datetime import datetime
        ano = datetime.now().year
    conn = get_db()
    try:
        MESES = {'Jan': 1, 'Fev': 2, 'Mar': 3, 'Abr': 4, 'Mai': 5, 'Jun': 6,
                 'Jul': 7, 'Ago': 8, 'Set': 9, 'Out': 10, 'Nov': 11, 'Dez': 12}
        
        q = "SELECT maquina, SUM(dur_min) as tempo_total_min, COUNT(*) as n_falhas FROM kpi_paradas_raw WHERE ano=?"
        vals = [ano]
        
        if semana in MESES:
            q += " AND mes=?"
            vals.append(MESES[semana])
        elif semana.startswith('S'):
            q += " AND semana_iso=?"
            vals.append(int(semana[1:]))
        else:
            q += " AND mes=?"
            vals.append(int(semana))
            
        if linha and linha != 'TODAS':
            q += " AND linha=?"
            vals.append(linha)
            
        q += " GROUP BY maquina ORDER BY tempo_total_min DESC"
        
        rows = conn.execute(q, vals).fetchall()
        
        # Calculate MTBF per machine
        result = []
        for r in rows:
            d = dict(r)
            maq = d['maquina']
            tp = d['tempo_total_min']
            n_falhas = d['n_falhas']
            
            # Find the line for this machine
            ln = linha
            if ln == 'TODAS':
                ln_row = conn.execute("SELECT linha FROM kpi_paradas_raw WHERE maquina=? AND linha != '' LIMIT 1", (maq,)).fetchone()
                ln = ln_row['linha'] if ln_row else 'Linha 4'
                
            # Fetch tempo disponivel
            td_q = "SELECT SUM(tempo_disponivel_min) as td FROM kpi_producao_raw WHERE linha=? AND ano=?"
            td_vals = [ln, ano]
            if semana in MESES:
                td_q += " AND mes=?"
                td_vals.append(MESES[semana])
            elif semana.startswith('S'):
                td_q += " AND semana_iso=?"
                td_vals.append(int(semana[1:]))
            else:
                td_q += " AND mes=?"
                td_vals.append(int(semana))
                
            prod = conn.execute(td_q, td_vals).fetchone()
            td = prod['td'] if (prod and prod['td'] and prod['td'] > 0) else (30 * 24 * 60 if semana in MESES else 7 * 24 * 60)
            
            mtbf = (td / n_falhas / 60) if n_falhas > 0 else (td / 60)
            d['mtbf_h'] = round(mtbf, 2)
            result.append(d)
            
        return result
    finally:
        conn.close()

@app.get("/api/kpi/metas-confiabilidade")
def get_metas_confiabilidade():
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM kpi_metas_confiabilidade ORDER BY linha").fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()

@app.post("/api/kpi/metas-confiabilidade")
async def save_metas_confiabilidade(req: Request):
    data = await req.json()  # lista de {linha, meta_mtbf_h, meta_mttr_h, meta_indisponibilidade_pct}
    conn = get_db()
    try:
        for item in data:
            conn.execute('''
                INSERT INTO kpi_metas_confiabilidade (linha, meta_mtbf_h, meta_mttr_h, meta_indisponibilidade_pct)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(linha) DO UPDATE SET
                    meta_mtbf_h=excluded.meta_mtbf_h,
                    meta_mttr_h=excluded.meta_mttr_h,
                    meta_indisponibilidade_pct=excluded.meta_indisponibilidade_pct
            ''', (item['linha'], item['meta_mtbf_h'], item['meta_mttr_h'], item['meta_indisponibilidade_pct']))
        conn.commit()
        return {"status": "ok"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.get("/api/prazo_ciente")
def get_prazo_ciente(email: str):
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")
    conn = get_db()
    rows = conn.execute("SELECT registro_id, faixa_prazo FROM registro_prazo_ciente WHERE user_email = ?", (email,)).fetchall()
    conn.close()
    return [{"registro_id": r["registro_id"], "faixa_prazo": r["faixa_prazo"]} for r in rows]

@app.post("/api/prazo_ciente")
async def post_prazo_ciente(req: Request):
    data = await req.json()
    email = data.get("email")
    notificacoes = data.get("notificacoes", [])
    if not email or not notificacoes:
        return {"status": "ok"}
    
    import datetime
    now_str = datetime.datetime.now().isoformat()
    
    conn = get_db()
    try:
        for n in notificacoes:
            conn.execute('''
                INSERT OR IGNORE INTO registro_prazo_ciente (user_email, registro_id, faixa_prazo, data_ciente)
                VALUES (?, ?, ?, ?)
            ''', (email, n["registro_id"], n["faixa_prazo"], now_str))
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
    
    return {"status": "ok"}

# ==========================================
# SSE (Server-Sent Events) - REALTIME SYNC
# ==========================================
@app.get("/api/stream")
async def sse_stream(request: Request):
    """
    Endpoint SSE que escuta as mudanças no arquivo SQLite.
    Como a arquitetura é distribuída (cada PC roda o seu server.py),
    o método mais rápido e leve de detectar que *outro* PC salvou algo 
    é checando a data de modificação (mtime) do arquivo database.sqlite.
    """
    async def event_generator():
        last_mtime = 0
        if os.path.exists(SYNC_PATH):
            last_mtime = os.path.getmtime(SYNC_PATH)
        
        while True:
            if await request.is_disconnected():
                break
            
            # Checa a data de modificação do arquivo de sincronização
            if os.path.exists(SYNC_PATH):
                current_mtime = os.path.getmtime(SYNC_PATH)
                if current_mtime > last_mtime:
                    last_mtime = current_mtime
                    # Arquivo de sync mudou! Envia um evento SSE para o frontend atualizar a tela
                    yield f"data: {json.dumps({'type': 'db_updated'})}\n\n"
            
            # Aguarda 1 segundo antes de checar novamente
            await asyncio.sleep(1)
            
    return StreamingResponse(event_generator(), media_type="text/event-stream")

@app.get("/")
def read_root():
    return FileResponse(os.path.join(ROOT_DIR, "index.html"))

app.mount("/", StaticFiles(directory=ROOT_DIR, html=True), name="static")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8081)
